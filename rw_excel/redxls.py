#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import xlrd
import xlwt
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import random


def showexcel(path):
    workbook = xlrd.open_workbook(path)
    sheets = workbook.sheet_names()
    worksheet = workbook.sheet_by_name(sheets[0])
    book = []
    for i in range(0, worksheet.nrows):
        row = worksheet.row(i)
        li = []
        for j in range(0, worksheet.ncols):
            sex = worksheet.cell_value(i, 1)
            address = worksheet.cell_value(i, 2)
            # study=worksheet.cell_value(i,3)
            if sex == '女' and address == '深圳':
                li.append(worksheet.cell_value(i, j))
                # print(worksheet.cell_value(i,j),"\t",end="")
            else:
                break
        if len(li) == 0:
            pass
        else:
            book.append(li)
    return book


def writeexcel03(read, path):
    if os.path.exists(path) == True:
        path = path + '/' + str(random.randint(12, 20)) + '.xls'
    wb = xlwt.Workbook()
    sheet = wb.add_sheet("ceshi")
    value = showexcel(read)  # 从一个文件里面读取数据
    # print(value)
    # value=[["name","python","php","js"],["price",'23','32','33'],['lan','china','english','china'],['time','2015','2012','2014']]
    for i in range(0, len(value)):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])
    wb.save(path)
    print("OK")


def writeexcel07(path):

    wb = Workbook()
    # sheet=wb.add_sheet("xlwt3数据测试表")
    sheet = wb.create_sheet(0, "xlwt3数据测试表")
    value = [["名称", "hadoop编程实战", "hbase编程实战", "lucene编程实战"], ["价格", "52.3", "45", "36"], [
        "出版社", "机械工业出版社", "人民邮电出版社", "华夏人民出版社"], ["中文版式", "中", "英", "英"]]
    sheet.cell(row=1, column=2).value = "温度"
    wb.save(path)
    print("写入数据成功！")


def read07excel(path):
    wb2 = load_workbook(path)

    ws = wb2.get_sheet_by_name("详单一")
    row = ws.get_highest_row()
    col = ws.get_highest_column()
    print("列数: ", ws.get_highest_column())
    print("行数: ", ws.get_highest_row())

    for i in range(0, row):
        for j in range(0, col):
            print(ws.rows[i][j].value, "\t\t", end="")

        print()

# x=showexcel('D:/item/test/test.xls')
# print(x)
writeexcel03('D:/item/test/test.xls', 'D:/item/test')
